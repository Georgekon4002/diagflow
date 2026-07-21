"""
DiagFlow — Mock Slis DB Seeder
==============================
Reads docs/init_data.xlsx and db/exam_codes.xlsx and
populates db/mock_slis.db (SQLite) using the schema in init.sql.

The init_data.xlsx contains real exam data.  This seeder
remaps all visitdates to the last 5 days (today through 4 days ago)
so the app's "3-day window" filter will return meaningful data.
Most rows have their diagnostician deliberately cleared so
the Pending tab has data to work with.

Usage:
    python db/seed_mock_db.py

Run from the project root (the diagflow/ workspace folder).
Re-running is safe — it drops and re-creates the tables each time.
"""

import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

# ── Locate files relative to this script ──────────────────────────
SCRIPT_DIR   = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DB_PATH      = SCRIPT_DIR / "mock_slis.db"
INIT_SQL     = SCRIPT_DIR / "init.sql"
DATA_XLSX    = PROJECT_ROOT / "docs" / "init_data.xlsx"   # initial exam data
CODES_XLSX   = PROJECT_ROOT / "docs" / "exam_codes.xlsx"  # exam category codes

try:
    import openpyxl
except ImportError:
    sys.exit(
        "ERROR: openpyxl is not installed.\n"
        "Install it with:  pip install openpyxl\n"
        "(or activate your venv first)"
    )


# ── Category mapping ──────────────────────────────────────────────
KATEGORYID_MAP = {
    18: "CT",
    22: "MRI",
    21: "MRA",
}


def _iso_date(val) -> str | None:
    """Convert an openpyxl date/datetime cell value to ISO-8601 string."""
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()
    s = str(val).strip()
    return s if s else None


def _int_or_none(val) -> int | None:
    """Return int or None; handles 'NULL' strings from the xlsx."""
    if val is None or str(val).strip().upper() == "NULL":
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _str_or_none(val) -> str | None:
    """Return stripped string or None; handles 'NULL' strings."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s.upper() == "NULL" or s == "" else s


# ── Load exam_codes.xlsx → {examnumcode: category} ───────────────
def load_exam_categories(path: Path) -> dict[int, tuple[str, str]]:
    """Returns {examnumcode: (name, category_str)}."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    result: dict[int, tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        code = _int_or_none(d.get("EXAMNUMCODE"))
        kid  = _int_or_none(d.get("KATEGORYID"))
        name = _str_or_none(d.get("NAME")) or ""
        if code is not None and kid in KATEGORYID_MAP:
            result[code] = (name, KATEGORYID_MAP[kid])
    print(f"  Loaded {len(result)} exam category mappings")
    return result


# ── Date remapping logic ───────────────────────────────────────────
def build_date_remap(source_dates: list) -> dict:
    """
    Maps the unique source visitdates to the last 5 days relative to today.

    Strategy:
      - Collect distinct source dates (oldest first)
      - Assign them to target dates: today, today-1, today-2, today-3, today-4
      - If there are more source dates than 5, cycle through the target dates
      - Result: ~all exams spread across the last 5 days
    """
    today = date.today()
    # Target dates: last 5 days, most recent first
    target_dates = [(today - timedelta(days=i)).isoformat() for i in range(5)]

    # Gather and sort unique source dates
    unique_src = sorted({d for d in source_dates if d}, reverse=True)  # newest first

    remap: dict[str, str] = {}
    for i, src in enumerate(unique_src):
        remap[src] = target_dates[i % len(target_dates)]

    # Print summary
    from collections import Counter
    tgt_counts = Counter(remap.values())
    print("  Date remap targets:")
    for d, n in sorted(tgt_counts.items(), reverse=True):
        print(f"    {d}: {n} source date(s) -> this target")

    return remap


# ── Seed the database ─────────────────────────────────────────────
def seed(db_path: Path, init_sql: Path, data_xlsx: Path, codes_xlsx: Path):
    print(f"Seeding {db_path} ...")

    # Load category map first
    print("Loading exam_codes.xlsx ...")
    cat_map = load_exam_categories(codes_xlsx)

    # Connect to SQLite
    con = sqlite3.connect(db_path)
    cur = con.cursor()

    # Drop existing tables so re-runs are idempotent
    cur.execute("DROP TABLE IF EXISTS slis_exams")
    cur.execute("DROP TABLE IF EXISTS exam_categories")
    cur.execute("DROP INDEX IF EXISTS idx_slis_exams_diagnostis")
    cur.execute("DROP INDEX IF EXISTS idx_slis_exams_visitdate")
    cur.execute("DROP INDEX IF EXISTS idx_slis_exams_examnumcode")
    con.commit()

    # Create schema directly (avoids SQL comment / semicolon parsing issues)
    print("Creating schema ...")
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS exam_categories (
            examnumcode   INTEGER PRIMARY KEY,
            name          TEXT    NOT NULL,
            category      TEXT    NOT NULL
        );

        CREATE TABLE IF NOT EXISTS slis_exams (
            oldexam         INTEGER,
            oldvisit        INTEGER,
            oldorder        TEXT,
            oldpers         INTEGER,
            olddiagnostis   TEXT,
            aa              INTEGER,
            extracode       INTEGER,          -- Order ID; NOT unique per exam
            visitid         INTEGER,
            demogid         INTEGER,
            fname           TEXT,
            lname           TEXT,
            examid          INTEGER,
            examnumcode     INTEGER,
            examname        TEXT,
            visitdate       TEXT,
            labcodeid       INTEGER,
            laboratoryname  TEXT,
            wardid          INTEGER,
            wcode           TEXT,
            wname           TEXT,
            diagnostis      INTEGER,
            personelid      INTEGER,
            code            TEXT,
            name            TEXT,
            notes           TEXT,
            exammoreid      INTEGER PRIMARY KEY,  -- Unique exam instance ID (TRUE PK)
            category        TEXT,
            slis_synced_at  TEXT DEFAULT NULL,
            FOREIGN KEY (examnumcode) REFERENCES exam_categories (examnumcode)
        );

        CREATE INDEX IF NOT EXISTS idx_slis_exams_diagnostis  ON slis_exams (diagnostis);
        CREATE INDEX IF NOT EXISTS idx_slis_exams_visitdate    ON slis_exams (visitdate);
        CREATE INDEX IF NOT EXISTS idx_slis_exams_extracode    ON slis_exams (extracode);
        CREATE INDEX IF NOT EXISTS idx_slis_exams_examnumcode  ON slis_exams (examnumcode);
    """)

    # Insert exam_categories
    print("Inserting exam_categories ...")
    cat_rows = [
        (code, name, category)
        for code, (name, category) in cat_map.items()
    ]
    cur.executemany(
        "INSERT OR REPLACE INTO exam_categories (examnumcode, name, category) VALUES (?, ?, ?)",
        cat_rows,
    )
    con.commit()
    print(f"  >> {len(cat_rows)} category rows inserted")

    # Load and process slis_exams from init_data.xlsx
    print(f"Loading {data_xlsx} ...")
    wb = openpyxl.load_workbook(data_xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # ── Pass 1: collect all raw rows and original visitdates ──────
    raw_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        extracode = _int_or_none(d.get("EXTRACODE"))
        if extracode is None:
            continue
        raw_rows.append(d)

    print(f"  Loaded {len(raw_rows)} rows from xlsx")

    # ── Build date remap ──────────────────────────────────────────
    source_dates = [_iso_date(d.get("VISITDATE")) for d in raw_rows]
    date_remap = build_date_remap(source_dates)

    # ── Pass 2: build insert rows with remapped dates ─────────────
    today = date.today()
    cutoff = (today - timedelta(days=3)).isoformat()

    # We want most rows (70%) to fall in the last 3 days (cutoff <= date <= today)
    # and some (30%) in days 4-5 to test the expiry window.
    # The date_remap already cycles through all 5 days, so just mark which ones
    # to keep their original diagnostician and which to clear.
    #
    # Rule:
    #   - If remapped date < cutoff (day 4 or 5): keep original diagnostician
    #     (these simulate "older" exams that were already assigned before the window)
    #   - If remapped date >= cutoff (last 3 days): clear diagnostician for ~85% of rows
    #     (these are the fresh unassigned ones the app should process)

    insert_sql = """
        INSERT OR REPLACE INTO slis_exams (
            oldexam, oldvisit, oldorder, oldpers, olddiagnostis,
            aa, extracode, visitid, demogid,
            fname, lname,
            examid, examnumcode, examname,
            visitdate, labcodeid, laboratoryname,
            wardid, wcode, wname,
            diagnostis, personelid, code, name,
            notes, exammoreid, category, slis_synced_at
        ) VALUES (
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?,
            ?, ?, ?,
            ?, ?, ?,
            ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?, NULL
        )
    """

    exam_rows = []
    kept_assigned = 0
    cleared = 0

    for i, d in enumerate(raw_rows):
        extracode   = _int_or_none(d.get("EXTRACODE"))
        examnumcode = _int_or_none(d.get("EXAMNUMCODE"))

        # Derive category from exam_codes.xlsx, fall back to CATEGORY column
        if examnumcode and examnumcode in cat_map:
            category = cat_map[examnumcode][1]
        else:
            raw_cat = _str_or_none(d.get("CATEGORY")) or ""
            if "ΑΞΟΝΙΚΗ" in raw_cat:
                category = "CT"
            elif "ΜΑΓΝΗΤΙΚΗ" in raw_cat:
                category = "MRI"
            else:
                category = raw_cat or None

        # OLDVISIT: the xlsx stores 0 for "never"; keep as integer
        oldvisit = _int_or_none(d.get("OLDVISIT"))
        if oldvisit is None:
            oldvisit = 0

        # OLDORDER: the xlsx uses 1900-01-01 as a sentinel for "no date"
        oldorder_raw = d.get("OLDORDER")
        oldorder = None
        if oldorder_raw:
            iso = _iso_date(oldorder_raw)
            if iso and iso != "1900-01-01":
                oldorder = iso

        # Remap visitdate to last 5 days
        original_visitdate = _iso_date(d.get("VISITDATE"))
        visitdate = date_remap.get(original_visitdate, today.isoformat())

        # Decide whether to keep or clear the diagnostician
        original_diagnostis = _int_or_none(d.get("DIAGNOSTIS"))
        if visitdate < cutoff:
            # Older than 3 days — keep original diagnostician (simulate old assigned)
            diagnostis  = original_diagnostis
            diag_code   = _str_or_none(d.get("CODE"))
            diag_name   = _str_or_none(d.get("NAME"))
            kept_assigned += 1
        else:
            # Within last 3 days — clear diagnostician for 85% of rows
            # Keep assigned for only 15% (every ~7th row that had one)
            if original_diagnostis and (i % 7 == 0):
                diagnostis = original_diagnostis
                diag_code  = _str_or_none(d.get("CODE"))
                diag_name  = _str_or_none(d.get("NAME"))
                kept_assigned += 1
            else:
                diagnostis = None
                diag_code  = None
                diag_name  = None
                cleared += 1

        exam_rows.append((
            _int_or_none(d.get("OLDEXAM")),
            oldvisit,
            oldorder,
            _int_or_none(d.get("OLDPERS")),
            _str_or_none(d.get("OLDDIAGNOSTIS")),

            _int_or_none(d.get("AA")),
            extracode,
            _int_or_none(d.get("VISITID")),
            _int_or_none(d.get("DEMOGID")),

            _str_or_none(d.get("FNAME")),
            _str_or_none(d.get("LNAME")),

            _int_or_none(d.get("EXAMID")),
            examnumcode,
            _str_or_none(d.get("EXAMNAME")),

            visitdate,
            _int_or_none(d.get("LABCODEID")),
            _str_or_none(d.get("LABORATORYNAME")),

            _int_or_none(d.get("WARDID")),
            _str_or_none(d.get("WCODE")),
            _str_or_none(d.get("WNAME")),

            diagnostis,
            _int_or_none(d.get("PERSONELID")),
            diag_code,
            diag_name,

            _str_or_none(d.get("NOTES")),
            _int_or_none(d.get("EXAMMOREID")),
            category,
        ))

    cur.executemany(insert_sql, exam_rows)
    con.commit()

    cur.execute("SELECT COUNT(*) FROM slis_exams")
    rows_inserted = cur.fetchone()[0]
    print(f"  >> {rows_inserted} unique exam rows in DB")

    # Summary
    cur.execute("SELECT COUNT(*) FROM slis_exams WHERE diagnostis IS NULL")
    pending = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM slis_exams WHERE diagnostis IS NOT NULL")
    assigned = cur.fetchone()[0]
    cur.execute("SELECT category, COUNT(*) FROM slis_exams GROUP BY category ORDER BY category")
    by_cat = cur.fetchall()
    cur.execute("SELECT visitdate, COUNT(*) FROM slis_exams GROUP BY visitdate ORDER BY visitdate DESC")
    by_date = cur.fetchall()

    print()
    print("-- Summary ------------------------------------------")
    print(f"  Total exams   : {rows_inserted}")
    print(f"  Pending       : {pending}  (diagnostician cleared)")
    print(f"  Kept assigned : {kept_assigned}  (older or sampled rows)")
    print(f"  By category   : {dict(by_cat)}")
    print(f"  By date       :")
    for d, c in by_date:
        marker = " <-- within 3-day window" if d and d >= cutoff else " <-- older (4-5 days)"
        print(f"    {d}: {c} exams{marker}")
    print(f"  DB written to : {db_path.resolve()}")
    print("-----------------------------------------------------")

    con.close()


if __name__ == "__main__":
    seed(DB_PATH, INIT_SQL, DATA_XLSX, CODES_XLSX)
